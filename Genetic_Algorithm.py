import itertools
import math
from dataclasses import dataclass

import numpy as np
import networkx as nx
import matplotlib.pyplot as plt
import random

from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
# from .utils import Get_Distance_Or_Flow
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components
from scipy.sparse import csr_array
from itertools import groupby
from collections import Counter
from networkx.algorithms import approximation
import sys
import xlrd
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook

from UI import open_file
from batch_topology import create_batch_topology, unique_values_in_list_of_lists
from production_performance import prod_efficiency

from variant_topology import config, topology, workstation
import networkx as nx
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from ttkbootstrap.constants import *
import pandas as pd

random.seed(1314141)


@dataclass
class chromosome:
    k_val: float
    iter_nr: int
    sequence: list


def run_GA():
    global open_filename

    file = askopenfile(mode='r', filetypes=[
        ('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])  # To open the file that you want.
    # ' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename=file.name)  # Load into openpyxl
    wb2 = wb.active

    sheets = wb.sheetnames
    sh1 = wb[sheets[0]]
    open_filename = sh1

    # print(wb.sheetnames)
    # print(sheets)

    row = sh1.max_row
    column = sh1.max_column

    batch_seq = [[] for i in range(column)]

    for i in range(1, column + 1):

        for j in range(2, row + 1):
            # print(sh1.cell(i,1).value)
            if sh1.cell(j, i).value != None:
                batch_seq[i - 1].append(sh1.cell(j, i).value)

    # batch_seq = [[1, 5, 9, 10, 2, 11, 13, 15, 7, 20],
    #              [1, 2, 7, 3, 5, 6, 8, 9, 13, 15, 19, 20],
    #              [1, 5, 8, 6, 3, 2, 4, 10, 15, 17, 20],
    #              [1, 8, 9, 10, 2, 11, 13, 15, 7, 20],
    #              [1, 4, 17, 3, 8, 9, 13, 15, 19, 20],
    #              [1, 6, 8, 6, 3, 12, 4, 10, 15, 17, 20],
    #              [1, 14, 8, 6, 13, 2, 4, 10, 15, 17, 20]]

    for i in batch_seq:
        print(i)

    init_population = []
    start_k = 1.2
    stop_k = 2.0
    step_k = 0.2

    ## Calculate 2 sets of population with different iteration value 25 and 35#####

    for i in range(int(start_k * 10), int(stop_k * 10), int(step_k * 10)):
        chrm_1 = chromosome(i / 10, 10, batch_seq)
        init_population.append(chrm_1)
        # print(i / 10)

    for i in range(int(start_k * 10), int(stop_k * 10), int(step_k * 10)):
        chrm_2 = chromosome(i / 10, 15, batch_seq)
        init_population.append(chrm_2)

    # for p in population:
    #     print(p)

    fitness_list = []
    topology_htable = dict()
    for i in range(len(init_population)):
        btop = create_batch_topology(init_population[i].sequence, i + 1, init_population[i].k_val,
                                     init_population[i].iter_nr)
        fitness_list.append(btop[0])
        topology_htable.update({btop[0]: (btop[1], btop[2],init_population[i].k_val,init_population[i].iter_nr)})
        # print(btop)

    print("Fitness list:", fitness_list)

    # for key, value in topology_htable.items():
    #     print(key, value)

    ########choosing the parents ####################
    middle_index = round(len(init_population) / 2)

    sorted_fitness1 = fitness_list[:middle_index]
    sorted_fitness2 = fitness_list[middle_index:]

    print(sorted_fitness1.index(sorted(sorted_fitness1)[0]))
    print(sorted_fitness2.index(sorted(sorted_fitness2)[0]))

    parent_1 = init_population[sorted_fitness1.index(sorted(sorted_fitness1)[0])]
    parent_2 = init_population[sorted_fitness2.index(sorted(sorted_fitness2)[0]) + middle_index]

    print("Parent 1:", parent_1)
    print("Parent 2:,", parent_2)

    ####3 Crossover part of Genetic algorithm##################
    off_population = []
    offspring_fitness = []

    offspring_1 = chromosome(parent_1.k_val, parent_2.iter_nr, parent_1.sequence)
    off_population.append(offspring_1)

    offspring_2 = chromosome(parent_2.k_val, parent_1.iter_nr, parent_2.sequence)
    off_population.append(offspring_2)

    for i in range(len(off_population)):
        off_top = create_batch_topology(off_population[i].sequence, i + 1, off_population[i].k_val,
                                        off_population[i].iter_nr)
        offspring_fitness.append(off_top[0])
        topology_htable.update({off_top[0]: (off_top[1],off_top[2],off_population[i].k_val,off_population[i].iter_nr)})
        # print("OFF spring topologies:", i + 1, off_top)

    print(min(offspring_fitness))

    # for key, value in topology_htable.items():
    #     print(key, value)

    ### Mutation function to be decided later#####3#
    mut_population = []
    mut_fitness = []

    ###Recursive operation until desired fitness achieved#############
    min_fitness = []

    def GA_recursion(itr1, itr2, rec_nr):
        new_population = []
        fit_list = []
        off_populn = []
        offspr_fitness = []
        start_k = 1.3
        stop_k = 2.0
        step_k = (stop_k - start_k) / 0.25
        print(f'The recursion number is {rec_nr} with iteration pari {itr1} and {itr2}')
        # print("Cleared length of population:", len(new_population))
        for i in range(int(start_k * 10), int(stop_k * 10), int(step_k)):
            chrm1 = chromosome(i / 10, itr1, batch_seq)
            new_population.append(chrm1)
            # print(i / 10)

        for i in range(int(start_k * 10), int(stop_k * 10), int(step_k)):
            chrm2 = chromosome(i / 10, itr2, batch_seq)
            new_population.append(chrm2)

        for i in range(len(new_population)):
            top = create_batch_topology(new_population[i].sequence, i + 1, new_population[i].k_val,
                                        new_population[i].iter_nr)
            fit_list.append(top[0])
            topology_htable.update({top[0]: (top[1], top[2],new_population[i].k_val,new_population[i].iter_nr)})

        print("The current population fitness list:", fit_list)

        m_index = round(len(new_population) / 2)

        sorted_fit1 = fit_list[:m_index]
        sorted_fit2 = fit_list[m_index:]

        p_1 = new_population[sorted_fit1.index(sorted(sorted_fit1)[0])]
        print("The index of 1st parent in fit list:", sorted_fit1.index(sorted(sorted_fit1)[0]))

        p_2 = new_population[sorted_fit2.index(sorted(sorted_fit2)[0]) + m_index]
        print("The index of 2nd parent in fit list:", sorted_fit2.index(sorted(sorted_fit2)[0]) + m_index)

        offspr_1 = chromosome(p_1.k_val, p_2.iter_nr, p_1.sequence)
        off_populn.append(offspr_1)
        offspr_2 = chromosome(p_2.k_val, p_1.iter_nr, p_2.sequence)
        off_populn.append(offspr_2)

        print("offspring 1 chromosome:", offspr_1)
        print("offspring 2 chromosome:", offspr_2)

        for i in range(len(off_populn)):
            otop = create_batch_topology(off_population[i].sequence, i + 1, off_population[i].k_val,
                                         off_population[i].iter_nr)
            offspr_fitness.append(otop[0])
            topology_htable.update({otop[0]: (otop[1], otop[2],off_population[i].k_val,off_population[i].iter_nr)})
            # print("OFF spring topologies:", i + 1, otop)

        print("fitness list of offspring in this iteration:", offspr_fitness)
        print("minimum fitnesss value in this iteration:", min(offspr_fitness))

        if min(fit_list) < min(offspr_fitness):
            gen_min_fitness = min(fit_list)
        else:
            gen_min_fitness = min(offspr_fitness)
        print("minimumfitness value in this generation:", gen_min_fitness)

        # print(int(step_k))
        # print(result)

        if min(offspr_fitness) > 500 and itr1 != 40 and itr2 != 45:
            min_fitness.append(gen_min_fitness)
            # new_population.clear()
            # fit_list.clear()
            # off_populn.clear()
            # offspr_fitness.clear()
            # del p_1
            # del p_2
            # del offspr_1
            # del offspr_2

            GA_recursion(itr1 + 5, itr2 + 5, rec_nr + 1)
            result = min(min_fitness)

        elif min(offspr_fitness) <= 500:
            result = min(offspr_fitness)

        else:
            result = 0

        return result

    ### Draw the optimal topology#####
    OptmialGraph = nx.MultiGraph()
    n_list = unique_values_in_list_of_lists(batch_seq)
    e_list = []
    for i in range(len(batch_seq)):
        for j in range(len(batch_seq[i]) - 1):
            # print(graph[i][j], graph[i][j+1])
            edges = [batch_seq[i][j], batch_seq[i][j + 1]]
            e_list.append(edges)

    OptmialGraph.add_nodes_from(n_list)
    OptmialGraph.add_edges_from(e_list)

    ##### END of GA ######
    if min(offspring_fitness) > 500:
        print("\n\nRecursion Started")
        final_fitness = (GA_recursion(20, 25, 1))
        print("the min fitness list:", min_fitness)
        print("The least possible fitness value:", final_fitness)
        print("The topology of the fittest value:", topology_htable[final_fitness])

        nx.draw(OptmialGraph, topology_htable[final_fitness][0], with_labels=True)
        plt.savefig('optimal topology found from GA recursion')
        plt.clf()

    elif min(offspring_fitness) <= 500:
        print("Fitness value found below 500:", min(offspring_fitness))
        nx.draw(OptmialGraph, topology_htable[min(offspring_fitness)][0], with_labels=True)
        plt.show('optimal topology found without GA recursion')
        plt.clf()

    width_dict = Counter(OptmialGraph.edges())
    edge_width = [[u, v, {'frequency': value}]
                  for ((u, v), value) in width_dict.items()]
    print("The frequency of edges", edge_width)

    "Production performance of the fittest solution"
    Qty_order = [10, 30, 50, 20, 60, 20, 40]
    #Qty_order = [100,100,100,100,100,100,100]
    #Qty_order = [1, 1, 1, 1, 1, 1, 1]
    print(prod_efficiency(batch_seq, topology_htable[final_fitness][0], Qty_order, topology_htable[final_fitness][1]))